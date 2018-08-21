import xml.etree.cElementTree as ET
from openpyxl import Workbook
import openpyxl

count = lambda ch : sum(el.isdigit() for el in ch) 
find_version = lambda ch : ch[ch.find('.')-2:ch.find('.')+9]

def extract_datum(ch):
	ls = ch.split()
	ll = [(elem, count(elem)) for elem in ls ]
	ch  = max(ll, key = lambda x: x[1] ) 
	return ch[0]

def main():
	wb = openpyxl.load_workbook("MS-TCU-AIVC_200701201821002098-D4a-ID08.xlsx")
	ws = wb['Check_List_to_fill']
	ch = ws['E69'].value
	IVClist = ET.Element("IVClist")
	brand = ET.SubElement(IVClist, "brand")
	fileId = ET.SubElement(IVClist, "fileId")
	RegisterList = ET.SubElement(IVClist, "RegisterList")
	Register = ET.SubElement(RegisterList, "Register")
	IVC_Id = ET.SubElement(Register, "IVC_Id")
	IVC_HW = ET.SubElement(Register, "IVC_HW")
	IVC_ref = ET.SubElement(Register, "IVC_ref")
	SW_version = ET.SubElement(Register, "SW_version")
	IMEI = ET.SubElement(Register, "IMEI")
	EID = ET.SubElement(Register, "EID")
	Profile = ET.SubElement(Register, "Profile")
	PemFile = ET.SubElement(Register, "PemFile")
	SKFiles = ET.SubElement(Register, "SKFiles")
	ICCID = ET.SubElement(Profile, "ICCID")
	IMSI = ET.SubElement(Profile, "IMSI")
	MSISDN = ET.SubElement(Profile, "MSISDN")
	Telco = ET.SubElement(Profile, "Telco")
	ls_elem = [brand, IVC_Id, IVC_HW, IVC_ref , IMEI, EID, ICCID, IMSI]
	ls_cells = ['F42','C2', 'E46', 'E45', 'E50', 'E53', 'E51', 'E52']
	fileId.text = ws['C3'].value+'_'+ws['C2'].value[14:]
	PemFile.text = ws['C2'].value+'.pem'
	SKFiles.text = ws['C2'].value+'.sk'
	SW_version.text =find_version( ws['E32'].value )
	MSISDN.text = '07493'
	Telco.text = 'Orange'
	for elem in zip(ls_elem, ls_cells): 
		elem[0].text = extract_datum(ws[elem[1]].value)
	tree = ET.ElementTree(IVClist)
	tree.write(SW_version.text+".xml")


if __name__ == "__main__": 
	main()
